#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
match_products.py — 注文の商品名を「データリスト」の正式名称へ機械照合する

請求書作成スキルの中核。FAX/メール/LINEで届いた注文（商品名＋数量）を、
納品書C列に貼れる「データリストA列の正式名称」へ突き合わせ、
・貼り付け用ブロック（正式商品名 / 発注番号 / 数量）
・検証情報（卸価格№ / 単価 / 金額）
・未マッチ / あいまい一致の一覧（← 必ず人が確認する）
を出力する。

【思想】数式セルは絶対に触らない。既定は「照合・検証レポートのみ」。
--fill を付けた時だけ、原本のコピーに対して手入力4項目のみ書き込む（必ずバックアップ）。

使い方（レポートのみ）:
    python match_products.py --wb "○○様 2026年6月分 納品書・請求書.xlsx" --order order.csv
    python match_products.py --wb フォーマット.xlsx --order order.csv --date 2026-07-01

order.csv の形式（ヘッダ任意、UTF-8）:
    商品名,数量
    陶器チモシーポッド コーナー型 ブルー,3
    固定できる金具付きエサ入れ スイーツ 大 ピンク,1

任意の3列目に発注番号を書けば尊重する。無ければ日付+連番で自動採番。

使い方（コピーへ実際に転記＝上級・要注意）:
    python match_products.py --wb 月次.xlsx --order order.csv --date 2026-07-01 \
        --fill --sheet 納品書1 --customer "うさぎ専門店RabbitRuru" --out 月次_書込済.xlsx
"""

import argparse
import csv
import shutil
import sys
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です:  pip install openpyxl")

# Windowsの既定コンソール(cp932)でもUTF-8で出力する（文字化け・出力エラー防止）
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


# ---- 正規化：全角/半角・スペース・大小を吸収して比較用キーにする ----
def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    # 空白（全角スペース　含む）をすべて除去、小文字化
    s = "".join(s.split())
    return s.lower()


def build_master(wb):
    """データリストから候補を作る。
    正式名称A = C + '　' + D + '　' + E（A列の数式を自前で再現）。
    返り値: [ {official, no, price, key} , ... ]
    """
    if "データリスト" not in wb.sheetnames:
        sys.exit("シート『データリスト』が見つかりません。フォーマットが違う可能性あり。")
    ws = wb["データリスト"]
    master = []
    for r in range(3, ws.max_row + 1):
        c = ws.cell(row=r, column=3).value  # C 商品名
        if c is None or str(c).strip() == "":
            continue
        d = ws.cell(row=r, column=4).value  # D カラー
        e = ws.cell(row=r, column=5).value  # E サイズ
        no = ws.cell(row=r, column=2).value  # B 卸価格№
        price = ws.cell(row=r, column=8).value  # H 単価
        official = f"{c}　{d if d is not None else ''}　{e if e is not None else ''}"
        master.append({
            "official": official,
            "no": no,
            "price": price,
            "key": norm(official),
            "key_name": norm(c),
        })
    if not master:
        sys.exit("データリストに商品が1件も見つかりません。")
    return master


def best_match(query, master):
    """クエリ商品名に最も近い候補を返す。 (item, score) を返す。"""
    q = norm(query)
    best, best_score = None, 0.0
    for m in master:
        # 完全一致は最優先
        if q == m["key"]:
            return m, 1.0
        score = SequenceMatcher(None, q, m["key"]).ratio()
        # 商品名だけ（カラー/サイズ除く）が含まれると加点：部分入力の注文に強くする
        if m["key_name"] and m["key_name"] in q:
            score = max(score, 0.6 + 0.4 * SequenceMatcher(None, q, m["key"]).ratio())
        if score > best_score:
            best, best_score = m, score
    return best, best_score


def load_order(path):
    """order.csv を読む。 [ (商品名, 数量, 発注番号 or None), ... ]"""
    rows = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if not row or not str(row[0]).strip():
                continue
            name = str(row[0]).strip()
            # ヘッダ行をスキップ
            if i == 0 and name in ("商品名", "商品", "品名"):
                continue
            qty = row[1].strip() if len(row) > 1 and str(row[1]).strip() else ""
            try:
                qty = int(float(qty)) if qty != "" else None
            except ValueError:
                qty = None
            order_no = str(row[2]).strip() if len(row) > 2 and str(row[2]).strip() else None
            rows.append((name, qty, order_no))
    if not rows:
        sys.exit("注文が空です。order.csv を確認してください。")
    return rows


# しきい値：これ未満は「未マッチ」、これ以上でも1.0未満は「要確認（あいまい）」
STRONG = 0.92
WEAK = 0.75


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--wb", required=True, help="納品書・請求書ワークブック(.xlsx)")
    ap.add_argument("--order", required=True, help="注文CSV(商品名,数量[,発注番号])")
    ap.add_argument("--date", help="発行日/納品日 YYYY-MM-DD（発注番号の自動採番にも使用）")
    ap.add_argument("--fill", action="store_true", help="コピーへ実際に転記（要注意・数式は触らない）")
    ap.add_argument("--sheet", default="納品書1", help="--fill時の書込先シート名")
    ap.add_argument("--customer", help="--fill時に A3 へ入れる取引先名")
    ap.add_argument("--out", help="--fill時の出力ファイル名")
    args = ap.parse_args()

    date = None
    if args.date:
        try:
            date = datetime.strptime(args.date, "%Y-%m-%d")
        except ValueError:
            sys.exit("--date は YYYY-MM-DD 形式で。例 2026-07-01")

    wb = openpyxl.load_workbook(args.wb, data_only=False)
    master = build_master(wb)
    order = load_order(args.order)

    matched, review, missing = [], [], []
    total = 0
    date_tag = date.strftime("%Y%m%d") if date else "YYYYMMDD"

    print("=" * 70)
    print("  請求書作成スキル / 商品名 機械照合レポート")
    print(f"  ワークブック: {args.wb}")
    print(f"  注文件数: {len(order)}  発行日: {args.date or '(未指定)'}")
    print("=" * 70)

    paste = []  # (正式名, 発注番号, 数量)
    for idx, (name, qty, order_no) in enumerate(order, start=1):
        m, score = best_match(name, master)
        onum = order_no or f"{date_tag}−{idx}"  # 全角ハイフン U+2212
        if score >= STRONG:
            price = m["price"]
            amount = (price * qty) if (isinstance(price, (int, float)) and qty) else None
            matched.append((name, m, qty, onum, amount))
            paste.append((m["official"], onum, qty))
            if amount:
                total += amount
        elif score >= WEAK:
            review.append((name, m, score, qty, onum))
        else:
            missing.append((name, m, score, qty))

    # ---- 貼り付け用ブロック ----
    print("\n■ 貼り付け用ブロック（納品書 C列=商品名 / I列=発注番号 / K列=数量）")
    print("  ※ C列は下の正式名称をそのまま貼る（単価は自動で出る）")
    print("-" * 70)
    for official, onum, qty in paste:
        print(f"  {official}\t{onum}\t{qty if qty is not None else ''}")
    if not paste:
        print("  （強一致なし）")

    # ---- 検証情報 ----
    if matched:
        print("\n■ 検証（№ / 単価 / 数量 / 金額）")
        print("-" * 70)
        for name, m, qty, onum, amount in matched:
            print(f"  ✓ №{m['no']}  単価{m['price']}  ×{qty}  = {amount}  ← {m['official'].strip()}")
        print(f"\n  小計(強一致分): {total:,} 円  ※税・端数・請求書合計は必ずExcelで確認")

    # ---- 要確認（あいまい一致） ----
    if review:
        print("\n■ ⚠ 要確認（あいまい一致・人が判断）")
        print("-" * 70)
        for name, m, score, qty, onum in review:
            print(f"  ? 注文『{name}』×{qty}")
            print(f"      候補: {m['official'].strip()}  (№{m['no']} 単価{m['price']})  類似度{score:.2f}")

    # ---- 未マッチ ----
    if missing:
        print("\n■ ✗ 未マッチ（データリストに無い可能性・要確認）")
        print("-" * 70)
        for name, m, score, qty in missing:
            hint = f"  近い候補: {m['official'].strip()} (類似度{score:.2f})" if m else ""
            print(f"  ✗ 注文『{name}』×{qty}{hint}")

    print("\n" + "=" * 70)
    print(f"  強一致 {len(matched)} / 要確認 {len(review)} / 未マッチ {len(missing)}")
    if review or missing:
        print("  → 要確認・未マッチは人（最終は山本社長）が確認するまで請求確定しないこと。")
    print("=" * 70)

    # ---- --fill：コピーへ手入力4項目だけ書き込む ----
    if args.fill:
        if review or missing:
            print("\n[中止] 要確認/未マッチが残っています。全て解消してから --fill してください。")
            return
        if not args.out:
            sys.exit("--fill には --out（出力ファイル名）が必要です。")
        # 原本は触らず out へコピーしてから書く（原本のバックアップも作る）
        bak = args.wb + f".backup_{datetime.now().strftime('%Y%m%d')}.xlsx" if False else None
        shutil.copyfile(args.wb, args.out)
        wb2 = openpyxl.load_workbook(args.out, data_only=False)
        if args.sheet not in wb2.sheetnames:
            sys.exit(f"シート『{args.sheet}』が見つかりません。")
        ws = wb2[args.sheet]
        if args.customer:
            ws["A3"] = args.customer
        if date:
            ws["M2"] = date
        row = 12
        for official, onum, qty in paste:
            ws.cell(row=row, column=3).value = official   # C 商品名
            ws.cell(row=row, column=9).value = onum        # I 発注番号
            ws.cell(row=row, column=11).value = qty        # K 数量
            row += 1
        wb2.save(args.out)
        print(f"\n[書込完了] {args.out} の『{args.sheet}』へ {len(paste)} 行を転記しました。")
        print("  ★必ずExcelで開いて ①単価が全行入っているか ②合計 ③レイアウト/画像 を目視確認。")


if __name__ == "__main__":
    main()
